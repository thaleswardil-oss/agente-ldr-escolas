import ast
import json
import math
import re
import time
import traceback
import unicodedata
from io import BytesIO
from datetime import datetime, timezone
from urllib.parse import urlparse

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from supabase import create_client
    ERRO_IMPORT_SUPABASE = ""
except Exception as exc:
    create_client = None
    ERRO_IMPORT_SUPABASE = f"{type(exc).__name__}: {exc}"

try:
    from tavily import TavilyClient
    ERRO_IMPORT_TAVILY = ""
except Exception as exc:
    TavilyClient = None
    ERRO_IMPORT_TAVILY = f"{type(exc).__name__}: {exc}"

st.set_page_config(page_title="Agente LDR Proesc v5", layout="wide")
st.markdown(
    """
    <style>
    h1, h2, h3, p { color: #004225 !important; }
    .stButton>button { background-color: #004225 !important; color: white !important; font-weight: bold; height: 3em; border-radius: 10px; }
    .stMetric { background-color: #f0f9eb; padding: 15px; border-radius: 10px; border: 1px solid #64CD32; }
    [data-testid="stExpander"] { border: 1px solid #64CD32; border-radius: 10px; }
    </style>
    """,
    unsafe_allow_html=True,
)

BUILD_ID = "2026.07.22.TAVILY.3"
VALOR_NAO_IDENTIFICADO = "Não identificado"
DELAYS_RETRY = [2, 4, 8, 16]
MAX_TENTATIVAS = len(DELAYS_RETRY) + 1
POLL_INTERVALO = 4
POLL_TIMEOUT = 300
PESOS = {
    "diretor": 25,
    "sge_atual": 25,
    "agenda_digital": 20,
    "telefone_alternativo": 15,
    "cnpj": 5,
    "email": 5,
    "site": 5,
}
PRIORIDADE_FONTES = {
    "site_oficial": 600,
    "receita": 500,
    "qsa": 400,
    "google_maps": 300,
    "instagram": 200,
    "facebook": 100,
    "outro": 0,
}
SIGLAS_PRESERVADAS = {
    "SESI", "SENAI", "COC", "SESC", "SENAC", "IF", "IFPA", "IFAP", "IFMA",
    "IFCE", "IFPI", "IFRN", "IFPE", "IFPB", "IFBA", "IFAL", "IFS", "IFES",
    "IFRJ", "IFSP", "IFPR", "IFSC", "IFRS", "IFC", "IFMS", "IFMT", "IFGO",
    "IFG", "IFTO", "UNICEF", "APAE", "CIEP", "EMEF", "EMEI", "EJA",
}
ARTIGOS_MINUSCULOS = {"da", "das", "de", "do", "dos", "e"}
MAPA_TERMOS = {
    "ESC": "Escola",
    "ESCOLA": "Escola",
    "COLEGIO": "Colégio",
    "COLÉGIO": "Colégio",
    "CENTRO": "Centro",
    "EDUCACIONAL": "Educacional",
    "EDUCACAO": "Educação",
    "EDUCAÇÃO": "Educação",
    "INFANTIL": "Infantil",
    "ENSINO": "Ensino",
    "POLICIA": "Polícia",
    "POLÍCIA": "Polícia",
    "MILITAR": "Militar",
    "GERACAO": "Geração",
    "GERAÇÃO": "Geração",
    "INTEGRADO": "Integrado",
    "INTEGRAL": "Integral",
    "TECNICO": "Técnico",
    "TÉCNICO": "Técnico",
    "TECNOLOGIA": "Tecnologia",
    "CRIANCA": "Criança",
    "CRIANÇA": "Criança",
    "JARDIM": "Jardim",
    "COOPERATIVA": "Cooperativa",
    "INSTITUTO": "Instituto",
    "FUNDACAO": "Fundação",
    "FUNDAÇÃO": "Fundação",
    "ASSOCIACAO": "Associação",
    "ASSOCIAÇÃO": "Associação",
}
VALORES_VAZIOS = {
    "", "null", "none", "nan", "n/a", "na", "não identificado",
    "nao identificado", "não informado", "nao informado", "não encontrado",
    "nao encontrado", "indisponível", "indisponivel", "sem dados",
    "desconhecido",
}
CAMPOS_RESULTADO = [
    "cnpj",
    "razao_social",
    "diretor",
    "telefone_alternativo",
    "email",
    "site",
    "sge_atual",
    "agenda_digital",
    "observacoes",
]


def agora_iso():
    return datetime.now(timezone.utc).isoformat()


def texto_seguro(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and math.isnan(valor):
        return ""
    return str(valor).strip()


def limitar_texto(valor, limite=12000):
    texto = texto_seguro(valor)
    if len(texto) <= limite:
        return texto
    return texto[:limite] + "…"


def valor_presente(valor):
    return texto_seguro(valor).lower() not in VALORES_VAZIOS


def normalizar_espacos(texto):
    return re.sub(r"\s+", " ", texto_seguro(texto)).strip()


def remover_acentos(texto):
    normalizado = unicodedata.normalize("NFD", texto_seguro(texto))
    return "".join(char for char in normalizado if unicodedata.category(char) != "Mn")


def serializar_json(valor):
    return json.dumps(valor, ensure_ascii=False, default=str)


def resumo_erro(excecao):
    return limitar_texto(f"{type(excecao).__name__}: {excecao}", 4000)


def objeto_para_dict(objeto):
    if objeto is None:
        return {}
    if isinstance(objeto, dict):
        return objeto
    if hasattr(objeto, "model_dump"):
        try:
            convertido = objeto.model_dump()
            if isinstance(convertido, dict):
                return convertido
        except Exception:
            pass
    if hasattr(objeto, "to_dict"):
        try:
            convertido = objeto.to_dict()
            if isinstance(convertido, dict):
                return convertido
        except Exception:
            pass
    if hasattr(objeto, "__dict__"):
        try:
            return {k: v for k, v in vars(objeto).items() if not k.startswith("_")}
        except Exception:
            pass
    return {}


def lista_segura(valor):
    if valor is None:
        return []
    if isinstance(valor, list):
        return valor
    if isinstance(valor, tuple):
        return list(valor)
    if isinstance(valor, set):
        return list(valor)
    return [valor]


def capitalizar_token(token, indice):
    bruto = texto_seguro(token)
    if not bruto:
        return ""
    superior = bruto.upper()
    sem_acento = remover_acentos(superior)
    if superior in SIGLAS_PRESERVADAS or sem_acento in SIGLAS_PRESERVADAS:
        return superior
    if sem_acento in MAPA_TERMOS:
        return MAPA_TERMOS[sem_acento]
    minusculo = bruto.lower()
    if indice > 0 and minusculo in ARTIGOS_MINUSCULOS:
        return minusculo
    if re.fullmatch(r"\d+[A-Za-z]?", bruto):
        return bruto.upper()
    if re.fullmatch(r"[IVXLCDM]+", superior):
        return superior
    return minusculo[:1].upper() + minusculo[1:]


def reorganizar_tipo_escola(tokens):
    if not tokens:
        return tokens
    for tipo in ["Colégio", "Escola"]:
        if tipo not in tokens:
            continue
        indice = tokens.index(tipo)
        if indice == 0:
            return tokens
        antes = tokens[:indice]
        depois = tokens[indice + 1:]
        conectivo = []
        if depois and depois[0].lower() in {"da", "das", "de", "do", "dos"}:
            conectivo = [depois[0].lower()]
            depois = depois[1:]
        return [tipo] + conectivo + antes + depois
    return tokens


def normalizar_nome(nome):
    bruto = normalizar_espacos(nome)
    if not bruto:
        return VALOR_NAO_IDENTIFICADO
    bruto = re.sub(r"[^\wÀ-ÿ\-\.\s]", " ", bruto, flags=re.UNICODE)
    tokens = []
    for token in normalizar_espacos(bruto).split(" "):
        if remover_acentos(token.upper()) in {"PART", "PARTICULAR"}:
            continue
        tokens.append(token)
    tratados = [capitalizar_token(token, indice) for indice, token in enumerate(tokens)]
    tratados = reorganizar_tipo_escola([token for token in tratados if token])
    tratados = [capitalizar_token(token, indice) for indice, token in enumerate(tratados)]
    return normalizar_espacos(" ".join(tratados)) or VALOR_NAO_IDENTIFICADO


def normalizar_municipio(municipio):
    bruto = normalizar_espacos(municipio)
    if not bruto:
        return VALOR_NAO_IDENTIFICADO
    return " ".join(capitalizar_token(token, indice) for indice, token in enumerate(bruto.split()))


def normalizar_uf(uf):
    texto = remover_acentos(texto_seguro(uf)).upper()
    return re.sub(r"[^A-Z]", "", texto)[:2]


def normalizar_cnpj(cnpj):
    digitos = re.sub(r"\D", "", texto_seguro(cnpj))
    if len(digitos) != 14 or digitos == digitos[0] * 14:
        return VALOR_NAO_IDENTIFICADO
    pesos_1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    soma_1 = sum(int(digitos[i]) * pesos_1[i] for i in range(12))
    resto_1 = soma_1 % 11
    dv_1 = 0 if resto_1 < 2 else 11 - resto_1
    pesos_2 = [6] + pesos_1
    soma_2 = sum(int(digitos[i]) * pesos_2[i] for i in range(13))
    resto_2 = soma_2 % 11
    dv_2 = 0 if resto_2 < 2 else 11 - resto_2
    if digitos[-2:] != f"{dv_1}{dv_2}":
        return VALOR_NAO_IDENTIFICADO
    return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"


def normalizar_email(email):
    texto = texto_seguro(email).lower()
    encontrados = re.findall(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", texto)
    return encontrados[0] if encontrados else VALOR_NAO_IDENTIFICADO


def normalizar_telefone(telefone):
    texto = texto_seguro(telefone)
    candidatos = re.findall(
        r"(?<!\d)(?:\+?55[\s.-]*)?(?:\(?([1-9]\d)\)?[\s.-]*)?(9?\d{4})[\s.-]?(\d{4})(?!\d)",
        texto,
    )
    ddds_validos = {
        "11", "12", "13", "14", "15", "16", "17", "18", "19", "21", "22", "24", "27", "28",
        "31", "32", "33", "34", "35", "37", "38", "41", "42", "43", "44", "45", "46", "47", "48", "49",
        "51", "53", "54", "55", "61", "62", "63", "64", "65", "66", "67", "68", "69", "71", "73", "74", "75",
        "77", "79", "81", "82", "83", "84", "85", "86", "87", "88", "89", "91", "92", "93", "94", "95", "96",
        "97", "98", "99",
    }
    for ddd, parte_1, parte_2 in candidatos:
        if ddd not in ddds_validos:
            continue
        numero = parte_1 + parte_2
        if len(numero) not in {8, 9} or len(set(numero)) == 1:
            continue
        if len(numero) == 9 and not numero.startswith("9"):
            continue
        return f"({ddd}) {numero[:5]}-{numero[5:]}" if len(numero) == 9 else f"({ddd}) {numero[:4]}-{numero[4:]}"
    return VALOR_NAO_IDENTIFICADO

def normalizar_url(url):
    texto = texto_seguro(url).strip(".,;:()[]{}<>\"'")
    if not texto:
        return VALOR_NAO_IDENTIFICADO
    if texto.startswith("www."):
        texto = "https://" + texto
    if not re.match(r"^https?://", texto, flags=re.IGNORECASE):
        if re.match(r"^[A-Za-z0-9.-]+\.[A-Za-z]{2,}(?:/.*)?$", texto):
            texto = "https://" + texto
        else:
            return VALOR_NAO_IDENTIFICADO
    try:
        parsed = urlparse(texto)
        if not parsed.netloc or "." not in parsed.netloc:
            return VALOR_NAO_IDENTIFICADO
        return texto
    except Exception:
        return VALOR_NAO_IDENTIFICADO


def normalizar_valor_generico(valor):
    texto = normalizar_espacos(valor)
    if texto.lower() in VALORES_VAZIOS:
        return VALOR_NAO_IDENTIFICADO
    return texto


def validar_diretor(valor):
    texto = normalizar_valor_generico(valor)
    if texto == VALOR_NAO_IDENTIFICADO:
        return VALOR_NAO_IDENTIFICADO
    termos_genericos = {
        "diretor", "diretora", "direção", "direcao", "gestor", "gestora",
        "coordenação", "coordenacao",
    }
    if texto.lower() in termos_genericos or len(texto) < 4:
        return VALOR_NAO_IDENTIFICADO
    return texto


def classificar_fonte(url, tipo_informado=""):
    tipo = remover_acentos(texto_seguro(tipo_informado)).lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "site": "site_oficial",
        "oficial": "site_oficial",
        "site_oficial": "site_oficial",
        "receita_federal": "receita",
        "receita": "receita",
        "qsa": "qsa",
        "quadro_societario": "qsa",
        "google_maps": "google_maps",
        "maps": "google_maps",
        "instagram": "instagram",
        "facebook": "facebook",
    }
    if tipo in aliases:
        return aliases[tipo]
    host = texto_seguro(url).lower()
    if "instagram.com" in host:
        return "instagram"
    if "facebook.com" in host or "fb.com" in host:
        return "facebook"
    if "google.com/maps" in host or "maps.google" in host or "goo.gl/maps" in host:
        return "google_maps"
    if "receita" in host or "gov.br/receitafederal" in host:
        return "receita"
    if "qsa" in host:
        return "qsa"
    return "outro"


def deduplicar_fontes(fontes):
    resultado = []
    vistos = set()
    for fonte in lista_segura(fontes):
        item = {"url": fonte} if isinstance(fonte, str) else objeto_para_dict(fonte)
        url = normalizar_url(item.get("url", item.get("uri", "")))
        if url == VALOR_NAO_IDENTIFICADO:
            continue
        chave = url.rstrip("/").lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        resultado.append(
            {
                "url": url,
                "titulo": limitar_texto(item.get("titulo", item.get("title", "")), 300),
                "tipo": classificar_fonte(url, item.get("tipo", item.get("tipo_fonte", ""))),
                "score": item.get("score", 0),
            }
        )
    return resultado


def limpar_bloco_json(texto):
    limpo = texto_seguro(texto)
    limpo = re.sub(r"^\s*```(?:json)?\s*", "", limpo, flags=re.IGNORECASE)
    limpo = re.sub(r"\s*```\s*$", "", limpo)
    limpo = limpo.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
    return limpo.strip()


def recuperar_json(valor):
    if isinstance(valor, dict):
        return valor
    if not isinstance(valor, str):
        convertido = objeto_para_dict(valor)
        if convertido:
            return convertido
        raise ValueError(f"Conteúdo estruturado inválido: {type(valor).__name__}")
    limpo = limpar_bloco_json(valor)
    erros = []
    for tentativa in [limpo, re.sub(r",\s*([}\]])", r"\1", limpo)]:
        try:
            convertido = json.loads(tentativa)
            if isinstance(convertido, dict):
                return convertido
        except Exception as exc:
            erros.append(resumo_erro(exc))
        decoder = json.JSONDecoder()
        for indice, char in enumerate(tentativa):
            if char not in "{[":
                continue
            try:
                convertido, _ = decoder.raw_decode(tentativa[indice:])
                if isinstance(convertido, dict):
                    return convertido
            except Exception:
                continue
    try:
        convertido = ast.literal_eval(limpo)
        if isinstance(convertido, dict):
            return convertido
    except Exception as exc:
        erros.append(resumo_erro(exc))
    raise ValueError(f"JSON inválido: {' | '.join(erros[-4:])}")


class PipelineLogger:
    def __init__(self, escola, callback=None):
        self.escola = escola
        self.callback = callback
        self.registros = []
        self.erros = []

    def registrar(self, step, mensagem, nivel="INFO", detalhe=""):
        registro = {
            "timestamp": agora_iso(),
            "escola": self.escola,
            "step": step,
            "nivel": nivel,
            "mensagem": limitar_texto(mensagem, 2000),
            "detalhe": limitar_texto(detalhe, 8000),
        }
        self.registros.append(registro)
        if nivel == "ERRO":
            self.erros.append(registro)
        if self.callback:
            try:
                self.callback(registro)
            except Exception as exc:
                self.registros.append(
                    {
                        "timestamp": agora_iso(),
                        "escola": self.escola,
                        "step": "LOG",
                        "nivel": "ERRO",
                        "mensagem": "Falha ao exibir log.",
                        "detalhe": resumo_erro(exc),
                    }
                )

    def info(self, step, mensagem, detalhe=""):
        self.registrar(step, mensagem, "INFO", detalhe)

    def alerta(self, step, mensagem, detalhe=""):
        self.registrar(step, mensagem, "ALERTA", detalhe)

    def erro(self, step, mensagem, excecao=None, detalhe=""):
        complemento = detalhe
        if excecao is not None:
            complemento = normalizar_espacos(f"{detalhe} {resumo_erro(excecao)}")
        self.registrar(step, mensagem, "ERRO", complemento)


def obter_segredo(nome, padrao=None, obrigatorio=False):
    try:
        valor = st.secrets[nome]
    except Exception:
        valor = padrao
    if obrigatorio and not texto_seguro(valor):
        raise RuntimeError(f"Secret obrigatório ausente: {nome}")
    return valor


@st.cache_resource(show_spinner=False)
def inicializar_servicos(supabase_url, supabase_key, tavily_api_key, tavily_project):
    if create_client is None:
        raise RuntimeError(f"supabase não pôde ser importado: {ERRO_IMPORT_SUPABASE}")
    if TavilyClient is None:
        raise RuntimeError(f"tavily-python não pôde ser importado: {ERRO_IMPORT_TAVILY}")
    supabase = create_client(supabase_url, supabase_key)
    kwargs = {"api_key": tavily_api_key}
    if texto_seguro(tavily_project):
        kwargs["project_id"] = tavily_project
    try:
        tavily = TavilyClient(**kwargs)
    except TypeError:
        tavily = TavilyClient(api_key=tavily_api_key)
    return supabase, tavily


def erro_transiente(excecao):
    texto = texto_seguro(excecao).lower()
    termos = [
        "429", "rate limit", "too many requests", "timeout", "timed out",
        "connection", "temporarily", "500", "502", "503", "504",
        "internal server", "service unavailable",
    ]
    return any(termo in texto for termo in termos)


def executar_com_retry(operacao, logger, step, descricao):
    ultimo_erro = None
    for tentativa in range(MAX_TENTATIVAS):
        logger.info(step, f"{descricao}: tentativa {tentativa + 1}/{MAX_TENTATIVAS}.")
        try:
            resposta = operacao()
            logger.info(step, f"{descricao}: resposta recebida.")
            return resposta
        except Exception as exc:
            ultimo_erro = exc
            logger.erro(step, f"{descricao}: falha na tentativa {tentativa + 1}.", exc)
            if tentativa >= len(DELAYS_RETRY) or not erro_transiente(exc):
                break
            espera = DELAYS_RETRY[tentativa]
            logger.alerta(step, f"Retry em {espera} segundos.")
            time.sleep(espera)
    raise RuntimeError(f"{descricao} falhou: {resumo_erro(ultimo_erro)}")


def chamada_search(tavily, query):
    parametros = {
        "query": query,
        "search_depth": "advanced",
        "topic": "general",
        "max_results": 10,
        "include_answer": "advanced",
        "include_raw_content": "markdown",
        "chunks_per_source": 5,
        "country": "brazil",
        "exclude_domains": ["johncatt.com", "querobolsa.com.br", "melhorescola.com.br", "escol.as", "escavador.com", "jusbrasil.com.br", "linkedin.com", "youtube.com", "pinterest.com", "exato.digital"],
    }
    try:
        return tavily.search(**parametros)
    except TypeError:
        parametros.pop("chunks_per_source", None)
        try:
            return tavily.search(**parametros)
        except TypeError:
            parametros["include_answer"] = True
            parametros["include_raw_content"] = True
            return tavily.search(**parametros)


def chamada_extract(tavily, urls, foco):
    parametros = {
        "urls": urls,
        "query": foco,
        "extract_depth": "advanced",
        "format": "markdown",
        "chunks_per_source": 5,
        "include_images": False,
    }
    try:
        return tavily.extract(**parametros)
    except TypeError:
        parametros.pop("format", None)
        parametros.pop("chunks_per_source", None)
        try:
            return tavily.extract(**parametros)
        except TypeError:
            parametros.pop("query", None)
            return tavily.extract(**parametros)


def schema_resultado_tavily():
    candidato = {
        "type": "object",
        "description": "Dado encontrado com prova verificável e URL da fonte.",
        "properties": {
            "valor": {"type": "string"},
            "tipo_fonte": {"type": "string", "enum": ["site_oficial", "receita", "qsa", "google_maps", "instagram", "facebook", "outro"]},
            "url": {"type": "string"},
            "evidencia": {"type": "string"},
            "confianca": {"type": "number"},
        },
        "required": ["valor", "tipo_fonte", "url", "evidencia", "confianca"],
    }
    lista = {"type": "array", "items": candidato}
    return {
        "properties": {
            "identidade_confirmada": {"type": "boolean"},
            "nome_oficial": {"type": "string"},
            "municipio_confirmado": {"type": "string"},
            "uf_confirmada": {"type": "string"},
            "justificativa_identidade": {"type": "string"},
            "razao_social_candidatos": lista,
            "cnpj_candidatos": lista,
            "site_candidatos": lista,
            "diretor_candidatos": lista,
            "socio_mantenedor_candidatos": lista,
            "telefone_alternativo_candidatos": lista,
            "email_candidatos": lista,
            "sge_atual_candidatos": lista,
            "agenda_digital_candidatos": lista,
            "observacoes": {"type": "string"},
            "observacoes_tecnologia": {"type": "string"},
            "conflitos": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["identidade_confirmada", "nome_oficial", "municipio_confirmado", "uf_confirmada", "justificativa_identidade", "razao_social_candidatos", "cnpj_candidatos", "site_candidatos", "diretor_candidatos", "socio_mantenedor_candidatos", "telefone_alternativo_candidatos", "email_candidatos", "sge_atual_candidatos", "agenda_digital_candidatos", "observacoes", "observacoes_tecnologia", "conflitos"],
    }

def chamada_research(tavily, prompt, modelo):
    parametros = {
        "input": prompt,
        "model": modelo,
        "stream": False,
        "output_schema": schema_resultado_tavily(),
    }
    try:
        return tavily.research(**parametros)
    except TypeError:
        parametros.pop("stream", None)
        return tavily.research(**parametros)


def aguardar_research(tavily, tarefa, logger):
    tarefa_dict = objeto_para_dict(tarefa)
    request_id = texto_seguro(tarefa_dict.get("request_id"))
    if not request_id:
        if tarefa_dict.get("content") is not None:
            return tarefa_dict
        raise ValueError(f"Tavily não retornou request_id: {limitar_texto(serializar_json(tarefa_dict), 3000)}")
    inicio = time.monotonic()
    logger.info("STEP 1", f"Research criado: {request_id}.")
    while time.monotonic() - inicio < POLL_TIMEOUT:
        resultado = tavily.get_research(request_id)
        resultado_dict = objeto_para_dict(resultado)
        status = texto_seguro(resultado_dict.get("status")).lower()
        logger.info("STEP 1", f"Status do Research: {status or 'desconhecido'}.")
        if status == "completed":
            return resultado_dict
        if status == "failed":
            raise RuntimeError(
                f"Research falhou: {limitar_texto(serializar_json(resultado_dict), 4000)}"
            )
        time.sleep(POLL_INTERVALO)
    raise TimeoutError(f"Research excedeu {POLL_TIMEOUT} segundos.")


def construir_query(dados, objetivo):
    nome, municipio, uf = dados["nome_escola"], dados["municipio"], dados["uf"]
    base = f'"{nome}" "{municipio}" "{uf}"'
    telefone = dados.get("telefone_inep", "")
    if valor_presente(telefone):
        base += f' "{telefone}"'
    return {
        "identidade": f'{base} site oficial endereço CNPJ razão social escola',
        "governanca": f'{base} diretor diretora mantenedor mantenedora sócio sócia QSA CNPJ',
        "contatos": f'{base} site oficial contato telefone WhatsApp email secretaria',
        "tecnologia": f'{base} site oficial portal do aluno login aplicativo agenda digital sistema de gestão escolar SGE ERP',
        "redes": f'{base} Instagram Facebook diretor diretora mantenedor escola oficial',
    }[objetivo]

def montar_evidencias_busca(respostas):
    blocos = []
    fontes = []
    for objetivo, resposta in respostas.items():
        resposta_dict = objeto_para_dict(resposta)
        answer = limitar_texto(resposta_dict.get("answer", ""), 8000)
        if answer:
            blocos.append(f"RESUMO {objetivo.upper()}\n{answer}")
        for item in lista_segura(resposta_dict.get("results", [])):
            resultado = objeto_para_dict(item)
            url = normalizar_url(resultado.get("url", ""))
            if url == VALOR_NAO_IDENTIFICADO:
                continue
            titulo = limitar_texto(resultado.get("title", ""), 500)
            conteudo = limitar_texto(
                resultado.get("raw_content") or resultado.get("content", ""),
                10000,
            )
            score = resultado.get("score", 0)
            fontes.append({"url": url, "titulo": titulo, "score": score})
            blocos.append(
                f"FONTE {objetivo.upper()}\nURL: {url}\nTÍTULO: {titulo}\nSCORE: {score}\nCONTEÚDO: {conteudo}"
            )
    return limitar_texto("\n\n".join(blocos), 60000), deduplicar_fontes(fontes)


def selecionar_urls_extracao(fontes, limite=15):
    ordenadas = sorted(
        fontes,
        key=lambda item: (
            PRIORIDADE_FONTES.get(item.get("tipo", "outro"), 0),
            float(item.get("score", 0) or 0),
        ),
        reverse=True,
    )
    return [item["url"] for item in ordenadas[:limite]]


def montar_evidencias_extract(resposta):
    resposta_dict = objeto_para_dict(resposta)
    blocos = []
    fontes = []
    for item in lista_segura(resposta_dict.get("results", [])):
        resultado = objeto_para_dict(item)
        url = normalizar_url(resultado.get("url", ""))
        if url == VALOR_NAO_IDENTIFICADO:
            continue
        conteudo = limitar_texto(resultado.get("raw_content", ""), 12000)
        fontes.append({"url": url, "titulo": "", "score": 0})
        blocos.append(f"EXTRAÇÃO\nURL: {url}\nCONTEÚDO: {conteudo}")
    falhas = []
    for item in lista_segura(resposta_dict.get("failed_results", [])):
        falhas.append(objeto_para_dict(item))
    return limitar_texto("\n\n".join(blocos), 50000), deduplicar_fontes(fontes), falhas


def construir_prompt_research(dados, evidencias):
    return f"""
Enriqueça UMA ÚNICA escola brasileira. A precisão é mais importante que preencher campos.

ESCOLA ALVO
Nome original no INEP: {dados['nome_original']}
Nome normalizado: {dados['nome_escola']}
Município obrigatório: {dados['municipio']}
UF obrigatória: {dados['uf']}
Telefone do INEP: {dados['telefone_inep']}

EVIDÊNCIAS COLETADAS
{evidencias}

REGRAS
1. Não misture escolas homônimas, unidades, cidades, estados ou mantenedoras.
2. Confirme município e UF antes de aceitar qualquer campo.
3. Confirme identidade apenas com nome compatível + município/UF + endereço, telefone, domínio oficial ou CNPJ.
4. Se a identidade não estiver confirmada, retorne todas as listas vazias.
5. Cada candidato precisa de URL completa e trecho textual que contenha ou comprove o valor.
6. Nunca use dados de terceiros, fornecedores, anúncios, diretórios internacionais ou escolas de outra localidade.
7. Nunca invente, complete ou deduza dados ausentes.
8. Busque diretor em site oficial, equipe, documentos institucionais e redes oficiais.
9. Busque sócio, administrador ou mantenedor em CNPJ/QSA e mantenha separado de diretor.
10. Procure SGE e agenda no site oficial, portal do aluno, área restrita, login, aplicativos, políticas, subdomínios, landing pages e comunicados.
11. Só informe SGE ou agenda com nome explícito, domínio do fornecedor, link de login/app ou documento oficial.
12. Google Classroom, WhatsApp, material didático, sistema de ensino e fornecedor financeiro não são SGE sem prova explícita.
13. Priorize site oficial, Receita/QSA, Google Maps, Instagram e Facebook.
14. Em conflito, mantenha candidatos separados e não escolha pela ordem de aparição.
"""

def preparar_candidatos(valor):
    if valor is None:
        return []
    if isinstance(valor, dict):
        if "valor" in valor:
            return [valor]
        return []
    if isinstance(valor, str):
        return [{"valor": valor, "tipo_fonte": "outro", "url": "", "confianca": 0}]
    candidatos = []
    for item in lista_segura(valor):
        if isinstance(item, dict):
            candidatos.append(item)
        elif isinstance(item, str):
            candidatos.append({"valor": item, "tipo_fonte": "outro", "url": "", "confianca": 0})
    return candidatos


def pontuar_candidato(candidato):
    item = objeto_para_dict(candidato)
    url = texto_seguro(item.get("url", ""))
    tipo = classificar_fonte(url, item.get("tipo_fonte", ""))
    try:
        confianca = float(item.get("confianca", 0))
    except Exception:
        confianca = 0
    confianca = max(0, min(100, confianca))
    return PRIORIDADE_FONTES.get(tipo, 0) * 1000 + confianca


def selecionar_candidato(candidatos, normalizador, logger, step, campo):
    validos = []
    for candidato in preparar_candidatos(candidatos):
        item = objeto_para_dict(candidato)
        normalizado = normalizador(item.get("valor", ""))
        if normalizado == VALOR_NAO_IDENTIFICADO:
            continue
        item["valor_normalizado"] = normalizado
        item["tipo_fonte_normalizado"] = classificar_fonte(
            item.get("url", ""), item.get("tipo_fonte", "")
        )
        item["pontuacao_fonte"] = pontuar_candidato(item)
        validos.append(item)
    if not validos:
        return VALOR_NAO_IDENTIFICADO, None
    validos.sort(key=lambda item: item["pontuacao_fonte"], reverse=True)
    valores_distintos = {item["valor_normalizado"].lower() for item in validos}
    if len(valores_distintos) > 1:
        logger.alerta(
            step,
            f"Conflito encontrado em {campo}; escolhida a fonte mais confiável.",
            serializar_json(validos[:5]),
        )
    return validos[0]["valor_normalizado"], validos[0]


def extrair_lista(objeto, campo):
    for chave in [campo, f"{campo}_candidatos", f"candidatos_{campo}"]:
        if chave in objeto:
            return objeto[chave]
    return []


def juntar_observacoes(*partes):
    resultado = []
    vistos = set()
    for parte in partes:
        texto = normalizar_espacos(parte)
        if not valor_presente(texto):
            continue
        chave = texto.lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        resultado.append(texto)
    return " | ".join(resultado) if resultado else VALOR_NAO_IDENTIFICADO


def step_0_normalizacao(nome, municipio, uf, telefone_inep, logger):
    logger.info("STEP 0", "Normalização iniciada.")
    resultado = {
        "nome_original": texto_seguro(nome),
        "nome_escola": normalizar_nome(nome),
        "municipio": normalizar_municipio(municipio),
        "uf": normalizar_uf(uf),
        "telefone_inep": normalizar_telefone(telefone_inep),
    }
    logger.info("STEP 0", f"Nome normalizado: {resultado['nome_escola']}.")
    return resultado


def step_1_pesquisa(dados, tavily, modelo_research, logger):
    logger.info("STEP 1", f"Buscando escola {dados['nome_escola']}.")
    respostas_busca = {}
    erros = []
    for objetivo in ["identidade", "governanca", "contatos", "tecnologia", "redes"]:
        query = construir_query(dados, objetivo)
        try:
            respostas_busca[objetivo] = executar_com_retry(
                lambda q=query: chamada_search(tavily, q),
                logger,
                "STEP 1",
                f"Tavily Search {objetivo}",
            )
        except Exception as exc:
            erros.append(f"Search {objetivo}: {resumo_erro(exc)}")
            logger.erro("STEP 1", f"Pesquisa {objetivo} falhou; pipeline continuará.", exc)
    evidencias_busca, fontes_busca = montar_evidencias_busca(respostas_busca)
    evidencias_extract = ""
    fontes_extract = []
    falhas_extract = []
    urls = selecionar_urls_extracao(fontes_busca)
    if urls:
        try:
            resposta_extract = executar_com_retry(
                lambda: chamada_extract(
                    tavily,
                    urls,
                    "Identidade da escola, CNPJ, razão social, diretor, telefone, email, SGE, portal do aluno e agenda digital",
                ),
                logger,
                "STEP 1",
                "Tavily Extract",
            )
            evidencias_extract, fontes_extract, falhas_extract = montar_evidencias_extract(resposta_extract)
            logger.info(
                "STEP 1",
                f"Extract concluído com {len(fontes_extract)} páginas e {len(falhas_extract)} falhas.",
            )
        except Exception as exc:
            erros.append(f"Extract: {resumo_erro(exc)}")
            logger.erro("STEP 1", "Tavily Extract falhou; pipeline continuará.", exc)
    evidencias = limitar_texto(
        "\n\n".join(parte for parte in [evidencias_busca, evidencias_extract] if parte),
        90000,
    )
    resultado_research = {}
    fontes_research = []
    request_id = ""
    try:
        tarefa = executar_com_retry(
            lambda: chamada_research(
                tavily,
                construir_prompt_research(dados, evidencias),
                modelo_research,
            ),
            logger,
            "STEP 1",
            "Tavily Research",
        )
        request_id = texto_seguro(objeto_para_dict(tarefa).get("request_id"))
        resposta_research = executar_com_retry(
            lambda: aguardar_research(tavily, tarefa, logger),
            logger,
            "STEP 1",
            "Polling Tavily Research",
        )
        conteudo = resposta_research.get("content")
        resultado_research = recuperar_json(conteudo)
        logger.info("STEP 1", "JSON estruturado recebido do Tavily Research.", limitar_texto(serializar_json(resultado_research), 5000))
        fontes_research = deduplicar_fontes(resposta_research.get("sources", []))
    except Exception as exc:
        erros.append(f"Research: {resumo_erro(exc)}")
        logger.erro("STEP 1", "Tavily Research falhou; será usado fallback determinístico.", exc)
    fontes = deduplicar_fontes(fontes_busca + fontes_extract + fontes_research)
    return {
        "evidencias": evidencias,
        "fontes": fontes,
        "research": resultado_research,
        "request_id": request_id,
        "modelo": modelo_research,
        "erros": erros,
        "pesquisa_sucesso": bool(evidencias or resultado_research),
        "falhas_extract": falhas_extract,
    }


def extrair_cnpjs_evidencias(evidencias):
    return []

def extrair_emails_evidencias(evidencias):
    return []

def extrair_telefones_evidencias(evidencias):
    return []

def fonte_esta_no_dossie(url, fontes):
    normalizada = normalizar_url(url)
    if normalizada == VALOR_NAO_IDENTIFICADO:
        return False
    dominio = urlparse(normalizada).netloc.lower().removeprefix("www.")
    for fonte in fontes:
        fonte_url = normalizar_url(objeto_para_dict(fonte).get("url", ""))
        if fonte_url == VALOR_NAO_IDENTIFICADO:
            continue
        outro = urlparse(fonte_url).netloc.lower().removeprefix("www.")
        if dominio == outro or dominio.endswith("." + outro) or outro.endswith("." + dominio):
            return True
    return False


def filtrar_candidatos_com_prova(candidatos, fontes, logger, step, campo):
    validos = []
    for candidato in preparar_candidatos(candidatos):
        item = objeto_para_dict(candidato)
        url = texto_seguro(item.get("url", ""))
        evidencia = normalizar_espacos(item.get("evidencia", ""))
        valor = normalizar_espacos(item.get("valor", ""))
        if not url or not evidencia or not valor:
            logger.alerta(step, f"Candidato de {campo} descartado por falta de URL, evidência ou valor.")
            continue
        if not fonte_esta_no_dossie(url, fontes):
            logger.alerta(step, f"Candidato de {campo} descartado: URL fora das fontes pesquisadas.", url)
            continue
        valor_limpo = re.sub(r"\W", "", remover_acentos(valor).lower())
        evidencia_limpa = re.sub(r"\W", "", remover_acentos(evidencia).lower())
        if len(valor_limpo) >= 5 and valor_limpo not in evidencia_limpa:
            logger.alerta(step, f"Candidato de {campo} descartado: trecho não comprova o valor.", serializar_json(item))
            continue
        validos.append(item)
    return validos


def identidade_compativel(dados, objeto):
    municipio = remover_acentos(texto_seguro(objeto.get("municipio_confirmado", ""))).lower()
    alvo = remover_acentos(dados["municipio"]).lower()
    return municipio == alvo and normalizar_uf(objeto.get("uf_confirmada", "")) == dados["uf"]


def step_2_validacao_identidade(dados, pesquisa, logger):
    logger.info("STEP 2", "Validação da identidade iniciada.")
    objeto = pesquisa.get("research", {})
    confirmada = objeto.get("identidade_confirmada", False)
    if isinstance(confirmada, str):
        confirmada = confirmada.strip().lower() in {"true", "sim", "yes", "1", "confirmada"}
    confirmada = bool(confirmada) and identidade_compativel(dados, objeto)
    if not confirmada:
        logger.erro("STEP 2", "Identidade não confirmada; nenhum dado externo será aceito.")
        return {"identidade_confirmada": False, "nome_oficial": dados["nome_escola"], "razao_social": VALOR_NAO_IDENTIFICADO, "cnpj": VALOR_NAO_IDENTIFICADO, "site": VALOR_NAO_IDENTIFICADO, "justificativa": normalizar_valor_generico(objeto.get("justificativa_identidade", "")), "conflitos": lista_segura(objeto.get("conflitos", [])), "selecoes": {}}
    fontes = pesquisa.get("fontes", [])
    razao = filtrar_candidatos_com_prova(extrair_lista(objeto, "razao_social"), fontes, logger, "STEP 2", "razão social")
    cnpjs = filtrar_candidatos_com_prova(extrair_lista(objeto, "cnpj"), fontes, logger, "STEP 2", "CNPJ")
    sites = filtrar_candidatos_com_prova(extrair_lista(objeto, "site"), fontes, logger, "STEP 2", "site")
    razao_social, razao_escolhida = selecionar_candidato(razao, normalizar_valor_generico, logger, "STEP 2", "razão social")
    cnpj, cnpj_escolhido = selecionar_candidato(cnpjs, normalizar_cnpj, logger, "STEP 2", "CNPJ")
    site, site_escolhido = selecionar_candidato(sites, normalizar_url, logger, "STEP 2", "site")
    logger.info("STEP 2", "Identidade confirmada e candidatos validados por fonte.")
    return {"identidade_confirmada": True, "nome_oficial": normalizar_nome(objeto.get("nome_oficial", dados["nome_escola"])), "razao_social": razao_social, "cnpj": cnpj, "site": site, "justificativa": normalizar_valor_generico(objeto.get("justificativa_identidade", "")), "conflitos": lista_segura(objeto.get("conflitos", [])), "selecoes": {"razao_social": razao_escolhida, "cnpj": cnpj_escolhido, "site": site_escolhido}}

def step_3_extracao(dados, pesquisa, identidade, logger):
    logger.info("STEP 3", "Extração de contatos iniciada.")
    if not identidade.get("identidade_confirmada"):
        return {"diretor": VALOR_NAO_IDENTIFICADO, "telefone_alternativo": VALOR_NAO_IDENTIFICADO, "email": VALOR_NAO_IDENTIFICADO, "observacoes": VALOR_NAO_IDENTIFICADO, "selecoes": {}}
    objeto, fontes = pesquisa.get("research", {}), pesquisa.get("fontes", [])
    diretores = filtrar_candidatos_com_prova(extrair_lista(objeto, "diretor"), fontes, logger, "STEP 3", "diretor")
    socios = filtrar_candidatos_com_prova(extrair_lista(objeto, "socio_mantenedor"), fontes, logger, "STEP 3", "sócio/mantenedor")
    telefones = filtrar_candidatos_com_prova(extrair_lista(objeto, "telefone_alternativo"), fontes, logger, "STEP 3", "telefone")
    emails = filtrar_candidatos_com_prova(extrair_lista(objeto, "email"), fontes, logger, "STEP 3", "e-mail")
    diretor, diretor_escolhido = selecionar_candidato(diretores, validar_diretor, logger, "STEP 3", "diretor")
    telefone, telefone_escolhido = selecionar_candidato(telefones, normalizar_telefone, logger, "STEP 3", "telefone")
    email, email_escolhido = selecionar_candidato(emails, normalizar_email, logger, "STEP 3", "e-mail")
    nomes_socios = [normalizar_valor_generico(x.get("valor", "")) for x in socios]
    nomes_socios = [x for x in nomes_socios if x != VALOR_NAO_IDENTIFICADO]
    observacoes = normalizar_valor_generico(objeto.get("observacoes", ""))
    if nomes_socios:
        observacoes = juntar_observacoes(observacoes, "Sócio/mantenedor identificado: " + "; ".join(dict.fromkeys(nomes_socios)))
    logger.info("STEP 3", "Extração concluída com validação de evidências.")
    return {"diretor": diretor, "telefone_alternativo": telefone, "email": email, "observacoes": observacoes, "selecoes": {"diretor": diretor_escolhido, "telefone_alternativo": telefone_escolhido, "email": email_escolhido, "socio_mantenedor": socios}}

def step_4_tecnologia(dados, pesquisa, identidade, extracao, logger):
    logger.info("STEP 4", "Análise de tecnologia iniciada.")
    if not identidade.get("identidade_confirmada"):
        return {"sge_atual": VALOR_NAO_IDENTIFICADO, "agenda_digital": VALOR_NAO_IDENTIFICADO, "observacoes_tecnologia": VALOR_NAO_IDENTIFICADO, "selecoes": {}}
    objeto, fontes = pesquisa.get("research", {}), pesquisa.get("fontes", [])
    sges = filtrar_candidatos_com_prova(extrair_lista(objeto, "sge_atual"), fontes, logger, "STEP 4", "SGE")
    agendas = filtrar_candidatos_com_prova(extrair_lista(objeto, "agenda_digital"), fontes, logger, "STEP 4", "Agenda Digital")
    sge, sge_escolhido = selecionar_candidato(sges, normalizar_valor_generico, logger, "STEP 4", "SGE")
    agenda, agenda_escolhida = selecionar_candidato(agendas, normalizar_valor_generico, logger, "STEP 4", "Agenda Digital")
    logger.info("STEP 4", "Análise de tecnologia concluída com validação de evidências.")
    return {"sge_atual": sge, "agenda_digital": agenda, "observacoes_tecnologia": normalizar_valor_generico(objeto.get("observacoes_tecnologia", "")), "selecoes": {"sge_atual": sge_escolhido, "agenda_digital": agenda_escolhida}}

def calcular_score_proesc(dados):
    return min(
        100,
        max(
            0,
            int(sum(peso for campo, peso in PESOS.items() if valor_presente(dados.get(campo)))),
        ),
    )


def definir_status(dados, score, houve_erro_fatal=False):
    if houve_erro_fatal and score == 0:
        return "Erro"
    if score == 0:
        return "Sem dados"
    tem_diretor = valor_presente(dados.get("diretor"))
    tem_contato = valor_presente(dados.get("telefone_alternativo")) or valor_presente(dados.get("email"))
    tem_tecnologia = valor_presente(dados.get("sge_atual")) or valor_presente(dados.get("agenda_digital"))
    if score >= 70 and tem_diretor and tem_contato and tem_tecnologia:
        return "Completa"
    return "Parcial"


def step_5_score(resultado, houve_erro_fatal, logger):
    logger.info("STEP 5", "Cálculo do score iniciado.")
    score = calcular_score_proesc(resultado)
    status = definir_status(resultado, score, houve_erro_fatal)
    logger.info("STEP 5", f"Score {score} e status {status}.")
    return score, status


def erro_coluna_inexistente(excecao):
    texto = texto_seguro(excecao)
    padroes = [
        r"Could not find the '([^']+)' column",
        r'column "([^"]+)" does not exist',
        r"column ([A-Za-z0-9_]+) does not exist",
        r"PGRST204.*?'([^']+)'",
    ]
    for padrao in padroes:
        match = re.search(padrao, texto, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1)
    return ""


def inserir_supabase_resiliente(tabela, payload, supabase, logger, step):
    payload_atual = dict(payload)
    removidas = []
    ultimo_erro = None
    tentativa = 0
    ajustes = 0
    limite_ajustes = len(payload_atual) + 5
    while tentativa < MAX_TENTATIVAS and ajustes <= limite_ajustes:
        try:
            resposta = supabase.table(tabela).insert(payload_atual).execute()
            dados = lista_segura(objeto_para_dict(resposta).get("data", []))
            logger.info(step, f"Salvo no banco na tentativa {tentativa + 1}.")
            return dados, removidas
        except Exception as exc:
            ultimo_erro = exc
            coluna = erro_coluna_inexistente(exc)
            if coluna and coluna in payload_atual:
                removidas.append(coluna)
                payload_atual.pop(coluna, None)
                ajustes += 1
                logger.alerta(step, f"Coluna indisponível no Supabase: {coluna}.")
                continue
            logger.erro(step, f"Erro no Supabase na tentativa {tentativa + 1}.", exc)
            if tentativa < len(DELAYS_RETRY) and erro_transiente(exc):
                espera = DELAYS_RETRY[tentativa]
                logger.alerta(step, f"Retry de banco em {espera} segundos.")
                time.sleep(espera)
            else:
                break
            tentativa += 1
    raise RuntimeError(f"Falha definitiva ao salvar em {tabela}: {resumo_erro(ultimo_erro)}")


def montar_observacoes_com_auditoria(observacoes, auditoria):
    resumo = {
        "tempo_gasto_segundos": auditoria.get("tempo_gasto_segundos"),
        "erro": auditoria.get("erro", ""),
        "ferramenta": "Tavily",
        "modelo_utilizado": auditoria.get("modelo_utilizado", ""),
        "request_id": auditoria.get("request_id", ""),
        "score": auditoria.get("score"),
        "status": auditoria.get("status"),
        "fontes": [fonte.get("url", "") for fonte in auditoria.get("fontes", []) if isinstance(fonte, dict)],
    }
    return juntar_observacoes(observacoes, f"AUDITORIA: {serializar_json(resumo)}")


def step_6_salvar(supabase, payload, auditoria, logger):
    logger.info("STEP 6", "Salvamento iniciado.")
    payload_completo = dict(payload)
    payload_completo["fontes"] = auditoria.get("fontes", [])
    payload_completo["tempo_gasto"] = auditoria.get("tempo_gasto_segundos", 0)
    payload_completo["erro"] = auditoria.get("erro", "")
    payload_completo["modelo_utilizado"] = auditoria.get("modelo_utilizado", "")
    payload_completo["auditoria"] = auditoria
    payload_completo["observacoes"] = montar_observacoes_com_auditoria(
        payload_completo.get("observacoes", ""),
        auditoria,
    )
    try:
        dados, removidas = inserir_supabase_resiliente(
            "leads_enriquecidos",
            payload_completo,
            supabase,
            logger,
            "STEP 6",
        )
        return {"salvo": True, "dados": dados, "colunas_removidas": removidas, "erro": ""}
    except Exception as exc:
        logger.erro("STEP 6", "Não foi possível salvar a escola.", exc)
        return {"salvo": False, "dados": [], "colunas_removidas": [], "erro": resumo_erro(exc)}


def pipeline_ldr(
    nome,
    municipio,
    uf,
    telefone_inep,
    rodada_id,
    supabase,
    tavily,
    modelo_research,
    callback_log=None,
):
    inicio = time.perf_counter()
    logger = PipelineLogger(normalizar_nome(nome), callback_log)
    erros_pipeline = []
    dados_norm = step_0_normalizacao(nome, municipio, uf, telefone_inep, logger)
    pesquisa = step_1_pesquisa(dados_norm, tavily, modelo_research, logger)
    erros_pipeline.extend(pesquisa.get("erros", []))
    houve_erro_fatal = not pesquisa.get("pesquisa_sucesso", False)
    identidade = step_2_validacao_identidade(dados_norm, pesquisa, logger)
    extracao = step_3_extracao(dados_norm, pesquisa, identidade, logger)
    tecnologia = step_4_tecnologia(dados_norm, pesquisa, identidade, extracao, logger)
    resultado = {
        "cnpj": identidade["cnpj"],
        "razao_social": identidade["razao_social"],
        "diretor": extracao["diretor"],
        "telefone_alternativo": extracao["telefone_alternativo"],
        "email": extracao["email"],
        "site": identidade["site"],
        "sge_atual": tecnologia["sge_atual"],
        "agenda_digital": tecnologia["agenda_digital"],
        "observacoes": juntar_observacoes(
            extracao["observacoes"],
            tecnologia["observacoes_tecnologia"],
            identidade["justificativa"] if not identidade["identidade_confirmada"] else "",
        ),
    }
    score, status = step_5_score(resultado, houve_erro_fatal, logger)
    tempo_gasto = round(time.perf_counter() - inicio, 2)
    erro_consolidado = " | ".join(dict.fromkeys(erro for erro in erros_pipeline if valor_presente(erro)))
    auditoria = {
        "fontes": pesquisa["fontes"],
        "tempo_gasto_segundos": tempo_gasto,
        "erro": limitar_texto(erro_consolidado, 10000),
        "ferramenta": "Tavily",
        "modelo_utilizado": pesquisa.get("modelo", ""),
        "request_id": pesquisa.get("request_id", ""),
        "score": score,
        "status": status,
        "identidade_confirmada": identidade["identidade_confirmada"],
        "conflitos": identidade.get("conflitos", []),
        "selecoes": {
            "identidade": identidade.get("selecoes", {}),
            "extracao": extracao.get("selecoes", {}),
            "tecnologia": tecnologia.get("selecoes", {}),
        },
        "falhas_extract": pesquisa.get("falhas_extract", []),
        "logs": logger.registros,
        "finalizado_em": agora_iso(),
        "build": BUILD_ID,
    }
    payload = {
        "rodada_id": rodada_id,
        "nome_escola": identidade["nome_oficial"] if valor_presente(identidade["nome_oficial"]) else dados_norm["nome_escola"],
        "municipio": dados_norm["municipio"],
        "uf": dados_norm["uf"],
        "telefone_inep": dados_norm["telefone_inep"],
        "status": status,
        "confianca": score,
        **resultado,
    }
    salvamento = step_6_salvar(supabase, payload, auditoria, logger)
    if not salvamento["salvo"]:
        auditoria["erro_salvamento"] = salvamento["erro"]
    return {
        "resultado": resultado,
        "payload": payload,
        "auditoria": auditoria,
        "salvamento": salvamento,
        "score": score,
        "status": status,
    }


def obter_celula(row, indice, padrao=""):
    try:
        return row.iloc[indice]
    except Exception:
        return padrao


def lista_opcoes_coluna(df, indice):
    if indice >= len(df.columns):
        return []
    valores = []
    for valor in df.iloc[:, indice].dropna().tolist():
        texto = texto_seguro(valor)
        if texto and texto not in valores:
            valores.append(texto)
    return sorted(valores)


def criar_rodada(supabase, nome_arquivo, total_leads, user_email):
    payload = {
        "nome_arquivo": limitar_texto(nome_arquivo or "Remessa Tavily", 255),
        "total_leads": int(total_leads),
        "usuario_email": user_email,
    }
    ultimo_erro = None
    for tentativa in range(MAX_TENTATIVAS):
        try:
            resposta = supabase.table("rodadas").insert(payload).execute()
            dados = lista_segura(objeto_para_dict(resposta).get("data", []))
            if not dados or "id" not in dados[0]:
                raise RuntimeError("Supabase não retornou o ID da rodada")
            return dados[0]["id"]
        except Exception as exc:
            ultimo_erro = exc
            if tentativa < len(DELAYS_RETRY) and erro_transiente(exc):
                time.sleep(DELAYS_RETRY[tentativa])
            else:
                break
    raise RuntimeError(f"Não foi possível criar a rodada: {resumo_erro(ultimo_erro)}")


def buscar_rodadas(supabase, user_email):
    resposta = (
        supabase.table("rodadas")
        .select("*")
        .eq("usuario_email", user_email)
        .order("created_at", desc=True)
        .execute()
    )
    return lista_segura(objeto_para_dict(resposta).get("data", []))


def buscar_leads_rodada(supabase, rodada_id):
    resposta = (
        supabase.table("leads_enriquecidos")
        .select("*")
        .eq("rodada_id", rodada_id)
        .order("confianca", desc=True)
        .execute()
    )
    return lista_segura(objeto_para_dict(resposta).get("data", []))


def valor_exportacao(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and math.isnan(valor):
        return ""
    if isinstance(valor, (dict, list, tuple, set)):
        return serializar_json(valor)
    texto = texto_seguro(valor)
    if texto.lower() in VALORES_VAZIOS:
        return "Não identificado"
    return texto


def limpar_observacoes_exportacao(valor):
    texto = valor_exportacao(valor)
    if "AUDITORIA:" in texto:
        texto = texto.split("AUDITORIA:", 1)[0].rstrip(" |")
    return texto or "Não identificado"


def preparar_dataframe_exportacao(leads):
    registros = []
    for lead in leads:
        item = objeto_para_dict(lead)
        registros.append(
            {
                "Escola": valor_exportacao(item.get("nome_escola")),
                "UF": valor_exportacao(item.get("uf")),
                "Município": valor_exportacao(item.get("municipio")),
                "Porte": valor_exportacao(item.get("porte", "")),
                "Endereço": valor_exportacao(item.get("endereco", item.get("endereço", ""))),
                "Tel. INEP": valor_exportacao(item.get("telefone_inep")),
                "CNPJ": valor_exportacao(item.get("cnpj")),
                "Razão Social": valor_exportacao(item.get("razao_social")),
                "Diretor / Responsável": valor_exportacao(
                    item.get("diretor", item.get("responsavel", item.get("socio_mantenedor", "")))
                ),
                "Tel. Alternativo": valor_exportacao(item.get("telefone_alternativo")),
                "E-mail": valor_exportacao(item.get("email")),
                "SGE Atual": valor_exportacao(item.get("sge_atual")),
                "Agenda Digital": valor_exportacao(item.get("agenda_digital")),
                "Site": valor_exportacao(item.get("site")),
                "Observações BDR": limpar_observacoes_exportacao(item.get("observacoes")),
            }
        )
    return pd.DataFrame(
        registros,
        columns=[
            "Escola",
            "UF",
            "Município",
            "Porte",
            "Endereço",
            "Tel. INEP",
            "CNPJ",
            "Razão Social",
            "Diretor / Responsável",
            "Tel. Alternativo",
            "E-mail",
            "SGE Atual",
            "Agenda Digital",
            "Site",
            "Observações BDR",
        ],
    )


def gerar_excel_resultados(leads):
    df_exportacao = preparar_dataframe_exportacao(leads)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Enriquecimento LDR"

    verde_escuro = "004225"
    verde_claro = "EAF5E4"
    borda_clara = "D9E5D3"
    branco = "FFFFFF"
    cinza_texto = "333333"
    borda = Border(
        left=Side(style="thin", color=borda_clara),
        right=Side(style="thin", color=borda_clara),
        top=Side(style="thin", color=borda_clara),
        bottom=Side(style="thin", color=borda_clara),
    )

    for coluna, cabecalho in enumerate(df_exportacao.columns, start=1):
        celula = worksheet.cell(row=1, column=coluna, value=cabecalho)
        celula.fill = PatternFill("solid", fgColor=verde_escuro)
        celula.font = Font(color=branco, bold=True, size=11)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda

    for linha, valores in enumerate(df_exportacao.itertuples(index=False, name=None), start=2):
        for coluna, valor in enumerate(valores, start=1):
            celula = worksheet.cell(row=linha, column=coluna, value=valor)
            celula.font = Font(color=cinza_texto, size=10)
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            celula.border = borda
            if linha % 2 == 0:
                celula.fill = PatternFill("solid", fgColor=verde_claro)
            if coluna in {6, 7, 10}:
                celula.number_format = "@"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:O{max(1, worksheet.max_row)}"
    worksheet.row_dimensions[1].height = 34
    for linha in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[linha].height = 48

    larguras = {
        "A": 30,
        "B": 8,
        "C": 22,
        "D": 14,
        "E": 36,
        "F": 18,
        "G": 20,
        "H": 32,
        "I": 28,
        "J": 18,
        "K": 30,
        "L": 22,
        "M": 22,
        "N": 34,
        "O": 60,
    }
    for coluna, largura in larguras.items():
        worksheet.column_dimensions[coluna].width = largura

    if worksheet.max_row >= 2:
        tabela = Table(displayName="TabelaEnriquecimentoLDR", ref=f"A1:O{worksheet.max_row}")
        estilo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        tabela.tableStyleInfo = estilo
        worksheet.add_table(tabela)

    worksheet.sheet_view.showGridLines = False
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), df_exportacao


st.title("🌿 Agente LDR Enterprise - Proesc v5")
st.sidebar.caption(f"Build ativo: {BUILD_ID}")

erro_inicializacao = ""
supabase = None
tavily = None
try:
    supabase_url = obter_segredo("SUPABASE_URL", obrigatorio=True)
    supabase_key = obter_segredo("SUPABASE_KEY", obrigatorio=True)
    tavily_api_key = obter_segredo("TAVILY_API_KEY", obrigatorio=True)
    tavily_project = obter_segredo("TAVILY_PROJECT_ID", "")
    supabase, tavily = inicializar_servicos(
        supabase_url,
        supabase_key,
        tavily_api_key,
        tavily_project,
    )
except Exception as exc:
    erro_inicializacao = resumo_erro(exc)

modelo_research = obter_segredo("TAVILY_RESEARCH_MODEL", "mini")
if modelo_research not in {"mini", "pro", "auto"}:
    modelo_research = "mini"

user_email = st.sidebar.text_input("E-mail BDR:", value="thales@proesc.com")
if not user_email:
    st.stop()

if erro_inicializacao:
    st.error(f"Erro de conexão: {erro_inicializacao}")
    st.stop()

tab1, tab2, tab3, tab4 = st.tabs(["📂 Upload", "🎯 Filtros", "🚀 Processar", "📜 Auditoria"])

with tab1:
    arquivo = st.file_uploader("Arraste sua planilha INEP aqui", type="xlsx")
    if arquivo:
        try:
            df_upload = pd.read_excel(arquivo)
            if df_upload.empty:
                st.error("A planilha está vazia.")
            elif len(df_upload.columns) < 13:
                st.error("A planilha não possui a estrutura mínima esperada do INEP.")
            else:
                st.session_state["df_raw"] = df_upload
                st.session_state["nome_arquivo"] = arquivo.name
                st.success("Planilha carregada!")
        except Exception as exc:
            st.error(f"Erro ao ler a planilha: {resumo_erro(exc)}")

with tab2:
    if "df_raw" in st.session_state:
        df = st.session_state["df_raw"].copy()
        try:
            if len(df.columns) > 14:
                mascara_processada = df.iloc[:, 14].astype(str).str.lower().str.contains(
                    r"sim|yes|true|1",
                    na=False,
                )
                df = df[~mascara_processada].copy()
            c1, c2, c3 = st.columns(3)
            opcoes_uf = lista_opcoes_coluna(df, 3)
            opcoes_porte = lista_opcoes_coluna(df, 12)
            with c1:
                f_uf = st.selectbox("UF", ["Todos"] + opcoes_uf)
            with c2:
                f_porte = st.selectbox("Porte", ["Todos"] + opcoes_porte)
            with c3:
                limite = st.number_input(
                    "Tamanho do Lote",
                    min_value=1,
                    max_value=100,
                    value=3,
                    step=1,
                )
            if f_uf != "Todos":
                df = df[df.iloc[:, 3].astype(str) == f_uf]
            if f_porte != "Todos":
                df = df[df.iloc[:, 12].astype(str) == f_porte]
            st.session_state["df_final"] = df.head(int(limite)).copy()
            st.metric("Escolas prontas", len(st.session_state["df_final"]))
        except Exception as exc:
            st.error(f"Erro ao aplicar os filtros: {resumo_erro(exc)}")
    else:
        st.info("Carregue uma planilha na aba Upload.")

with tab3:
    if "df_final" in st.session_state:
        df_p = st.session_state["df_final"]
        if df_p.empty:
            st.warning("Nenhuma escola disponível com os filtros atuais.")
        if st.button("🚀 INICIAR CICLO DE INTELIGÊNCIA", disabled=df_p.empty):
            try:
                rid = criar_rodada(
                    supabase,
                    st.session_state.get("nome_arquivo", "Remessa Tavily"),
                    len(df_p),
                    user_email,
                )
            except Exception as exc:
                st.error(resumo_erro(exc))
                rid = None
            if rid is not None:
                progresso = st.progress(0)
                status_txt = st.empty()
                log_placeholder = st.empty()
                linhas_log = []
                sucessos = 0
                falhas_salvamento = 0

                def atualizar_log(registro):
                    horario = registro["timestamp"][11:19]
                    linha = f"{horario} | {registro['nivel']} | {registro['step']} | {registro['mensagem']}"
                    if registro.get("detalhe"):
                        linha += f" | {registro['detalhe']}"
                    linhas_log.append(linha)
                    log_placeholder.code("\n".join(linhas_log[-250:]), language="text")

                for posicao, (_, row) in enumerate(df_p.iterrows(), start=1):
                    nome = obter_celula(row, 1)
                    municipio = obter_celula(row, 4)
                    uf = obter_celula(row, 3)
                    telefone_inep = obter_celula(row, 8)
                    nome_exibicao = normalizar_nome(nome)
                    status_txt.info(f"Analisando: **{nome_exibicao}**")
                    try:
                        retorno = pipeline_ldr(
                            nome=nome,
                            municipio=municipio,
                            uf=uf,
                            telefone_inep=telefone_inep,
                            rodada_id=rid,
                            supabase=supabase,
                            tavily=tavily,
                            modelo_research=modelo_research,
                            callback_log=atualizar_log,
                        )
                        if retorno["salvamento"]["salvo"]:
                            sucessos += 1
                        else:
                            falhas_salvamento += 1
                            st.error(f"{nome_exibicao}: {retorno['salvamento']['erro']}")
                    except Exception as exc:
                        falhas_salvamento += 1
                        detalhe = limitar_texto(traceback.format_exc(), 8000)
                        atualizar_log(
                            {
                                "timestamp": agora_iso(),
                                "nivel": "ERRO",
                                "step": "PIPELINE",
                                "mensagem": f"Erro não tratado em {nome_exibicao}; próxima escola será processada.",
                                "detalhe": detalhe,
                            }
                        )
                        payload_erro = {
                            "rodada_id": rid,
                            "nome_escola": nome_exibicao,
                            "municipio": normalizar_municipio(municipio),
                            "uf": normalizar_uf(uf),
                            "telefone_inep": normalizar_telefone(telefone_inep),
                            "status": "Erro",
                            "confianca": 0,
                            "cnpj": VALOR_NAO_IDENTIFICADO,
                            "razao_social": VALOR_NAO_IDENTIFICADO,
                            "diretor": VALOR_NAO_IDENTIFICADO,
                            "telefone_alternativo": VALOR_NAO_IDENTIFICADO,
                            "email": VALOR_NAO_IDENTIFICADO,
                            "site": VALOR_NAO_IDENTIFICADO,
                            "sge_atual": VALOR_NAO_IDENTIFICADO,
                            "agenda_digital": VALOR_NAO_IDENTIFICADO,
                            "observacoes": f"Erro não tratado no pipeline: {resumo_erro(exc)}",
                        }
                        auditoria_erro = {
                            "fontes": [],
                            "tempo_gasto_segundos": 0,
                            "erro": resumo_erro(exc),
                            "ferramenta": "Tavily",
                            "modelo_utilizado": modelo_research,
                            "request_id": "",
                            "score": 0,
                            "status": "Erro",
                            "logs": linhas_log[-50:],
                            "traceback": detalhe,
                            "finalizado_em": agora_iso(),
                            "build": BUILD_ID,
                        }
                        logger_fallback = PipelineLogger(nome_exibicao, atualizar_log)
                        salvamento_erro = step_6_salvar(
                            supabase,
                            payload_erro,
                            auditoria_erro,
                            logger_fallback,
                        )
                        if salvamento_erro["salvo"]:
                            sucessos += 1
                            falhas_salvamento -= 1
                        else:
                            st.error(
                                f"{nome_exibicao}: erro no pipeline e no salvamento: {salvamento_erro['erro']}"
                            )
                    progresso.progress(posicao / len(df_p))
                status_txt.empty()
                if falhas_salvamento == 0:
                    st.success(f"✅ Rodada finalizada! {sucessos} escolas salvas.")
                else:
                    st.warning(
                        f"Rodada finalizada com {sucessos} escolas salvas e {falhas_salvamento} falhas de salvamento."
                    )
    else:
        st.info("Configure o lote na aba Filtros.")

with tab4:
    st.subheader("Auditoria de Rodadas")
    try:
        rodadas_db = buscar_rodadas(supabase, user_email)
        if not rodadas_db:
            st.info("Nenhuma rodada encontrada para este usuário.")
        for rodada in rodadas_db:
            criado_em = texto_seguro(rodada.get("created_at", ""))
            data_exibicao = criado_em[:16] if criado_em else "Sem data"
            total_leads = rodada.get("total_leads", 0)
            with st.expander(f"📁 {data_exibicao} | {total_leads} leads"):
                try:
                    leads_db = buscar_leads_rodada(supabase, rodada.get("id"))
                    if leads_db:
                        excel_bytes, df_exportacao = gerar_excel_resultados(leads_db)
                        st.dataframe(df_exportacao, use_container_width=True, hide_index=True)
                        st.download_button(
                            "📥 Baixar Excel",
                            excel_bytes,
                            f"remessa_ldr_proesc_{data_exibicao[:10]}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{rodada.get('id')}",
                        )
                    else:
                        st.info("Nenhum lead salvo nesta rodada.")
                except Exception as exc:
                    st.error(f"Erro ao carregar os leads da rodada: {resumo_erro(exc)}")
    except Exception as exc:
        st.error(f"Erro ao carregar auditoria: {resumo_erro(exc)}")
